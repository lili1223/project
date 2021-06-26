# -*- coding: utf-8 -*-
# @Time : 2021/5/23 15:59
# @Author : 11858
# @Email : w1111@qq.com
# @File : do_excel.py
# @Project : Project
from openpyxl import load_workbook
from Tools.project_path import *
from conf.read_config import ReadConfig
class ReadExcel:
    def __init__(self,file_name):
        self.file_name=file_name
    def get_data(self):
        mode=eval(ReadConfig.get_config(config_path,'MODE','mode'))
        user=820573
        test_data = []
        wb=load_workbook(self.file_name)
        for key in mode:
            sheet=wb[key]
            if mode[key] == 'all':
                for i in range(2,sheet.max_row+1):
                    sub_data={}
                    sub_data['case_id']=sheet.cell(i,1).value
                    sub_data['title'] = sheet.cell(i, 2).value
                    sub_data['url'] = sheet.cell(i, 3).value
                    if sheet.cell(i, 4).value.find("${user}") != -1:
                        sub_data['data'] = sheet.cell(i, 4).value.replace("${user}",str(user))
                    elif sheet.cell(i, 4).value.find("${user_1}") != -1:
                        sub_data['data'] = sheet.cell(i, 4).value.replace("${user_1}", str(user+1))
                    else:
                        sub_data['data'] = sheet.cell(i, 4).value
                    sub_data['method'] = sheet.cell(i, 5).value
                    sub_data['expected'] = sheet.cell(i, 6).value
                    sub_data['sheet_name'] = key
                    test_data.append(sub_data)
            else:
                for case_id in mode[key]:
                    sub_data = {}
                    sub_data['case_id'] = sheet.cell(case_id + 1, 1).value
                    sub_data['title'] = sheet.cell(case_id + 1, 2).value
                    sub_data['url'] = sheet.cell(case_id + 1, 3).value
                    sub_data['data'] = sheet.cell(case_id + 1, 4).value
                    sub_data['method'] = sheet.cell(case_id + 1, 5).value
                    sub_data['expected'] = sheet.cell(case_id + 1, 6).value
                    sub_data['sheet_name'] = key
                    test_data.append(sub_data)
        return test_data
    def wrire_back(self,sheet_name,i,res,result):
        wb=load_workbook(self.file_name)
        sheet=wb[sheet_name]
        sheet.cell(i, 7).value=res
        sheet.cell(i, 8).value=result
        wb.save(self.file_name)
if __name__ == '__main__':
    test=ReadExcel(case_data_path).get_data()
    # print(len(test)
    print(test)

