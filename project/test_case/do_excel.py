# -*- coding: utf-8 -*-
# @Time : 2021/5/23 15:59
# @Author : 11858
# @Email : w1111@qq.com
# @File : do_excel.py
# @Project : Project
from openpyxl import load_workbook
from Tools.project_path import *
from conf.read_config import ReadConfig
class ReadExcel:
    def __init__(self,file_name,sheet_name):
        self.file_name=file_name
        self.sheet_name=sheet_name
    def get_data(self):
        mode=eval(ReadConfig.get_config(config_path,'MODE','mode'))
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet_name]
        test_data=[]
        for i in range(2,sheet.max_row+1):
            sub_data={}
            sub_data['case_id']=sheet.cell(i,1).value
            sub_data['title'] = sheet.cell(i, 2).value
            sub_data['url'] = sheet.cell(i, 3).value
            sub_data['data'] = sheet.cell(i, 4).value
            sub_data['method'] = sheet.cell(i, 5).value
            sub_data['expected'] = sheet.cell(i, 6).value
            test_data.append(sub_data)
            if mode=='all':
                final_data=test_data
            else:
                final_data=[]
                for item in test_data:
                    if item['case_id'] in mode:
                        final_data.append(item)
        return test_data

    def wrire_back(self,i,res,result):
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet_name]
        sheet.cell(i, 7).value=res
        sheet.cell(i, 8).value=result
        wb.save(self.file_name)
if __name__ == '__main__':
    test=ReadExcel(case_data_path,'create_coupon').get_data()
    print(test)

